"""
Excel Exporter for KERBERUS Tabular Review.

Exports review data to Excel with:
- Main data sheet with all extracted values
- Citations sheet with source references
- Summary sheet with review metadata
"""

import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
from io import BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export review data to Excel format.
    
    Creates a workbook with:
    - Sheet 1: Review Data (main table)
    - Sheet 2: Citations (source references)
    - Sheet 3: Summary (metadata)
    """
    
    def __init__(self):
        """Initialize exporter."""
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check that openpyxl is available."""
        try:
            import openpyxl
            self._has_openpyxl = True
        except ImportError:
            logger.warning("openpyxl not installed - Excel export disabled")
            self._has_openpyxl = False
    
    def export_review(
        self,
        review,  # Review object
        output_path: Optional[str] = None
    ) -> BytesIO:
        """
        Export a review to Excel.
        
        Args:
            review: Review object
            output_path: Optional file path to save (if None, returns BytesIO)
            
        Returns:
            BytesIO with Excel file content
        """
        if not self._has_openpyxl:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Import preset to get field info
        from .presets import get_preset
        preset = get_preset(review.preset_id)
        
        wb = Workbook()
        
        # =====================================================================
        # Sheet 1: Review Data
        # =====================================================================
        ws_data = wb.active
        ws_data.title = "Review Data"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["#", "Document"] + [f.display_name for f in preset.fields if f.name != "document_name"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, row in enumerate(review.rows, start=2):
            # Index
            ws_data.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            
            # Document name
            ws_data.cell(row=row_idx, column=2, value=row.filename).border = thin_border
            
            # Field values
            col_idx = 3
            for field_def in preset.fields:
                if field_def.name == "document_name":
                    continue
                    
                field_data = row.fields.get(field_def.name, {})
                value = field_data.get("value")
                
                # Format value
                if value is None:
                    display_value = ""
                elif isinstance(value, bool):
                    display_value = "Yes" if value else "No"
                elif isinstance(value, list):
                    display_value = ", ".join(str(v) for v in value)
                else:
                    display_value = str(value)
                
                cell = ws_data.cell(row=row_idx, column=col_idx, value=display_value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                col_idx += 1
        
        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            # Set minimum width based on header, max 50
            width = min(max(len(header) + 2, 12), 50)
            ws_data.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws_data.freeze_panes = "A2"
        
        # =====================================================================
        # Sheet 2: Citations
        # =====================================================================
        ws_citations = wb.create_sheet("Citations")
        
        # Headers
        citation_headers = ["#", "Document", "Field", "Value", "Page", "Section", "Source Quote"]
        for col_idx, header in enumerate(citation_headers, start=1):
            cell = ws_citations.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Citation data
        citation_row = 2
        for row_idx, row in enumerate(review.rows, start=1):
            for field_def in preset.fields:
                field_data = row.fields.get(field_def.name, {})
                citation = field_data.get("citation")
                
                if citation and citation.get("quote"):
                    ws_citations.cell(row=citation_row, column=1, value=row_idx).border = thin_border
                    ws_citations.cell(row=citation_row, column=2, value=row.filename).border = thin_border
                    ws_citations.cell(row=citation_row, column=3, value=field_def.display_name).border = thin_border
                    ws_citations.cell(row=citation_row, column=4, value=str(field_data.get("value", ""))).border = thin_border
                    ws_citations.cell(row=citation_row, column=5, value=citation.get("page", "")).border = thin_border
                    ws_citations.cell(row=citation_row, column=6, value=citation.get("section", "")).border = thin_border
                    
                    quote_cell = ws_citations.cell(row=citation_row, column=7, value=citation.get("quote", ""))
                    quote_cell.alignment = Alignment(wrap_text=True)
                    quote_cell.border = thin_border
                    
                    citation_row += 1
        
        # Set column widths for citations
        citation_widths = [5, 25, 20, 25, 8, 12, 60]
        for col_idx, width in enumerate(citation_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws_citations.column_dimensions[col_letter].width = width
        
        ws_citations.freeze_panes = "A2"
        
        # =====================================================================
        # Sheet 3: Summary
        # =====================================================================
        ws_summary = wb.create_sheet("Summary")
        
        summary_data = [
            ("Review Name", review.name),
            ("Review ID", review.review_id),
            ("Preset", review.preset_name),
            ("Document Count", review.document_count),
            ("Status", review.status),
            ("Created", review.created_at),
            ("Last Updated", review.updated_at),
            ("Exported", datetime.utcnow().isoformat()),
        ]
        
        ws_summary.cell(row=1, column=1, value="Property").font = Font(bold=True)
        ws_summary.cell(row=1, column=2, value="Value").font = Font(bold=True)
        
        for row_idx, (prop, value) in enumerate(summary_data, start=2):
            ws_summary.cell(row=row_idx, column=1, value=prop)
            ws_summary.cell(row=row_idx, column=2, value=str(value))
        
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 50
        
        # =====================================================================
        # Save
        # =====================================================================
        if output_path:
            wb.save(output_path)
            logger.info(f"Exported review to {output_path}")
        
        # Return as BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_filename(self, review) -> str:
        """Generate a filename for the export."""
        # Clean review name for filename
        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in review.name)
        safe_name = safe_name.replace(" ", "_")[:50]
        
        date_str = datetime.now().strftime("%Y%m%d")
        
        return f"{safe_name}_{date_str}.xlsx"
